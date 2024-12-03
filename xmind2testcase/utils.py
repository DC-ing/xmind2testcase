#!/usr/bin/env python
# _*_ coding:utf-8 _*_
import json
import os
import xmind
import logging
import arrow
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Alignment, Border, Side

from xmind2testcase.parser import xmind_to_testsuites, config_sep
from xmind2testcase.config import TestResult


def get_absolute_path(path):
    """
    Return the absolute path of a file

    If path contains a start point (eg Unix '/') then use the specified start point
    instead of the current working directory. The starting point of the file path is
    allowed to begin with a tilde "~", which will be replaced with the user's home directory.
    """
    fp, fn = os.path.split(path)
    if not fp:
        fp = os.getcwd()
    fp = os.path.abspath(os.path.expanduser(fp))
    return os.path.join(fp, fn)


def get_xmind_testsuites(xmind_file):
    """Load the XMind file and parse to `xmind2testcase.metadata.TestSuite` list"""
    xmind_file = get_absolute_path(xmind_file)
    workbook = xmind.load(xmind_file)
    xmind_content_dict = workbook.getData()
    logging.debug(
        "loading XMind file(%s) dict data: %s", xmind_file, xmind_content_dict
    )

    if xmind_content_dict:
        testsuites = xmind_to_testsuites(xmind_content_dict)
        return testsuites
    else:
        logging.error("Invalid XMind file(%s): it is empty!", xmind_file)
        return []


def time_to_chinese(time_str):
    # 使用arrow库获取相对时间字符串
    relative_time_str = arrow.get(time_str).humanize()

    # 定义正则表达式模式和替换规则
    patterns = [
        (re.compile(r"(\d+)\s*seconds\s*ago"), lambda m: f"{int(m.group(1))}秒前"),
        (re.compile(r"(\d+)\s*minutes\s*ago"), lambda m: f"{int(m.group(1))}分钟前"),
        (re.compile(r"(\d+)\s*hours\s*ago"), lambda m: f"{int(m.group(1))}小时前"),
        (re.compile(r"(\d+)\s*days\s*ago"), lambda m: f"{int(m.group(1))}天前"),
        (re.compile(r"just\s*now"), lambda m: "刚刚"),
        # 如果需要处理未来的时间，可以添加以下模式
        # (re.compile(r'in\s*(\d+)\s*seconds'), lambda m: f"{int(m.group(1))}秒后"),
        # (re.compile(r'in\s*(\d+)\s*minutes'), lambda m: f"{int(m.group(1))}分钟后"),
        # (re.compile(r'in\s*(\d+)\s*hours'), lambda m: f"{int(m.group(1))}小时后"),
        # (re.compile(r'in\s*(\d+)\s*days'), lambda m: f"{int(m.group(1))}天后"),
    ]

    # 遍历模式列表并替换匹配项
    for pattern, replacement in patterns:
        relative_time_str = pattern.sub(replacement, relative_time_str)

    return relative_time_str


def get_xmind_testsuite_list(xmind_file):
    """Load the XMind file and get all testsuite in it

    :param xmind_file: the target XMind file
    :return: a list of testsuite data
    """
    xmind_file = get_absolute_path(xmind_file)
    logging.info(
        "Start converting XMind file(%s) to testsuite data list...", xmind_file
    )
    testsuite_list = get_xmind_testsuites(xmind_file)
    suite_data_list = []

    for testsuite in testsuite_list:
        product_statistics = {
            "case_num": 0,
            "non_execution": 0,
            "pass": 0,
            "failed": 0,
            "blocked": 0,
            "skipped": 0,
        }
        for sub_suite in testsuite.sub_suites:
            suite_statistics = {
                "case_num": len(sub_suite.testcase_list),
                "non_execution": 0,
                "pass": 0,
                "failed": 0,
                "blocked": 0,
                "skipped": 0,
            }
            for case in sub_suite.testcase_list:
                if case.result == TestResult.NON_EXECUTION.val:
                    suite_statistics["non_execution"] += 1
                elif case.result == TestResult.PASS.val:
                    suite_statistics["pass"] += 1
                elif case.result == TestResult.FAILED.val:
                    suite_statistics["failed"] += 1
                elif case.result == TestResult.BLOCKED.val:
                    suite_statistics["blocked"] += 1
                elif case.result == 4:
                    suite_statistics["skipped"] += 1
                else:
                    logging.warning(
                        "This testcase result is abnormal: %s, please check it: %s",
                        case.result,
                        case.to_dict(),
                    )
            sub_suite.statistics = suite_statistics
            for item in product_statistics:
                product_statistics[item] += suite_statistics[item]

        testsuite.statistics = product_statistics
        suite_data = testsuite.to_dict()
        suite_data_list.append(suite_data)

    logging.info(
        "Convert XMind file(%s) to testsuite data list successfully!", xmind_file
    )
    return suite_data_list


def get_xmind_testcase_list(xmind_file):
    """Load the XMind file and get all testcase in it

    :param xmind_file: the target XMind file
    :return: a list of testcase data
    """
    xmind_file = get_absolute_path(xmind_file)
    logging.info(
        "Start converting XMind file(%s) to testcases dict data...", xmind_file
    )
    testsuites = get_xmind_testsuites(xmind_file)
    testcases = []

    # 一个画布一个测试集
    for testsuite in testsuites:
        product = testsuite.name
        for suite in testsuite.sub_suites:
            for case in suite.testcase_list:
                case_data = case.to_dict()
                sub_suite_names = case_data["sub_suite"].split(config_sep["sep"])
                case_data["product"] = product
                case_data["suite"] = suite.name.strip()
                case_data["second_suite"] = sub_suite_names[0].strip()
                case_data["third_suite"] = (
                    sub_suite_names[1].strip() if len(sub_suite_names) >= 2 else "-"
                )
                # 优先使用用例上的执行人
                if case_data["writer"] == "":
                    case_data["writer"] = suite.writer if suite.writer else ""
                # 转化测试结果
                case_data["result"] = TestResult.get_desc(case_data["result"])
                # if case_data["result"] == TestResult.default_desc():
                #     case_data["result"] = ""
                testcases.append(case_data)
    logging.info(
        "Convert XMind file(%s) to testcases dict data successfully!", xmind_file
    )
    return testcases


def xmind_testsuite_to_json_file(xmind_file):
    """Convert XMind file to a testsuite json file"""
    xmind_file = get_absolute_path(xmind_file)
    logging.info(
        "Start converting XMind file(%s) to testsuites json file...", xmind_file
    )
    testsuites = get_xmind_testsuite_list(xmind_file)
    testsuite_json_file = xmind_file[:-6] + "_testsuite.json"

    if os.path.exists(testsuite_json_file):
        os.remove(testsuite_json_file)
        # logging.info('The testsuite json file already exists, return it directly: %s', testsuite_json_file)
        # return testsuite_json_file

    with open(testsuite_json_file, "w", encoding="utf8") as f:
        f.write(
            json.dumps(testsuites, indent=4, separators=(",", ": "), ensure_ascii=False)
        )
        logging.info(
            "Convert XMind file(%s) to a testsuite json file(%s) successfully!",
            xmind_file,
            testsuite_json_file,
        )

    return testsuite_json_file


def xmind_testcase_to_json_file(xmind_file):
    """Convert XMind file to a testcase json file"""
    xmind_file = get_absolute_path(xmind_file)
    logging.info(
        "Start converting XMind file(%s) to testcases json file...", xmind_file
    )
    testcases = get_xmind_testcase_list(xmind_file)
    testcase_json_file = xmind_file[:-6] + ".json"

    if os.path.exists(testcase_json_file):
        os.remove(testcase_json_file)
        # logging.info('The testcase json file already exists, return it directly: %s', testcase_json_file)
        # return testcase_json_file

    with open(testcase_json_file, "w", encoding="utf8") as f:
        f.write(
            json.dumps(testcases, indent=4, separators=(",", ": "), ensure_ascii=False)
        )
        logging.info(
            "Convert XMind file(%s) to a testcase json file(%s) successfully!",
            xmind_file,
            testcase_json_file,
        )

    return testcase_json_file


def dict_list_to_excel(
    dict_list,
    file_name,
    dropdown_fields=[],
    dropdown_options={},
    min_width=10,
    max_width=30,
    merge_fields=[],
):
    # 将列表转化为 DataFrame
    df = pd.DataFrame(dict_list)

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 绿色填充（RGB颜色代码）
    header_fill = PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid"
    )

    # 定义边框样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 将 DataFrame 的表头添加至工作表，并应用绿色填充和边框
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            vertical="center", wrap_text=True
        )  # 上下居中并自动换行

    # 将 DataFrame 转换为工作表，并为数据单元格添加边框
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center", wrap_text=True
            )  # 上下居中并自动换行

    # 初始下拉菜单数据验证集合
    dv_collection = {}

    for col_name in dropdown_fields:
        col_idx = list(df.columns).index(col_name) + 1
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(dropdown_options[col_name])}"',
            allow_blank=True,
        )
        dv.error = "输入值非法"
        dv.errorTitle = "非法输入"
        dv.prompt = "请选择一个选项"
        dv.promptTitle = "下拉选项"

        # 遍历整列，对每个单元格应用数据验证
        for row in range(2, len(df) + 2):
            cell = ws.cell(row=row, column=col_idx)
            dv.add(cell)

        dv_collection[col_name] = dv

    # 将数据验证集合添加至工作表
    for col_name, dv in dv_collection.items():
        ws.add_data_validation(dv)

    # 自动调整列宽并设置单元格自动换行
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = max(
            min_width, min(max_length + 2, max_width)
        )  # 适应的宽度在指定范围内
        ws.column_dimensions[column].width = adjusted_width

    # 合并指定列中连续相同的单元格
    for col_name in merge_fields:
        col_idx = list(df.columns).index(col_name) + 1
        start_row = 2
        current_value = None
        for row in range(2, len(df) + 2):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value != current_value:
                if row > start_row + 1:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=col_idx,
                        end_row=row - 1,
                        end_column=col_idx,
                    )
                start_row = row
                current_value = cell_value
        # 处理最后一组连续相同的单元格
        if len(df) + 1 > start_row:
            ws.merge_cells(
                start_row=start_row,
                start_column=col_idx,
                end_row=len(df) + 1,
                end_column=col_idx,
            )

    # 保存工作簿至文件
    wb.save(file_name)
